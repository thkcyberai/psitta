#!/usr/bin/env python3
"""Phase D Excel writer for the psitta-ad-search skill.

Reads samples/phase_c_results.json (230 enriched ads), writes a polished
20-column workbook to the skill root for the marketing agent to consume.

  Input:  samples/phase_c_results.json
  Output: psitta_ad_research_<YYYYMMDD>.xlsx

Sort: gold first, silver after; within each tier descending by heat.
Header row is frozen, auto-filtered, bold white on dark fill. Enrichment
text columns wrap. preview_image cells are hyperlinks. tier cells get a
gold/silver fill for at-a-glance scanning.

Consumes zero AdLibrary credits — pure file transform.
"""
from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_PATH = Path(__file__).resolve()
SCRIPTS_DIR = SCRIPT_PATH.parent
SKILL_DIR = SCRIPTS_DIR.parent
INPUT_PATH = SKILL_DIR / "samples" / "phase_c_results.json"
OUTPUT_NAME = f"psitta_ad_research_{datetime.now().strftime('%Y%m%d')}.xlsx"
OUTPUT_PATH = SKILL_DIR / OUTPUT_NAME

# ── Column schema ───────────────────────────────────────────────────────────
# (header, width, kind) where kind ∈ {"text", "int", "wrap", "url", "tier"}
COLUMNS = [
    ("tier",                      8,  "tier"),
    ("advertiser",                18, "text"),
    ("platform",                  18, "text"),
    ("days_running",              12, "int"),
    ("heat",                      12, "int"),
    ("impression",                12, "int"),
    ("est_spend_usd",             12, "int"),
    ("video_duration_s",          8,  "int"),
    ("cta",                       18, "text"),
    ("source_keyword",            22, "text"),
    ("source_bucket",             18, "text"),
    ("keyword_overlap",           8,  "int"),
    ("ads_type",                  8,  "int"),
    ("ads_promote_type",          18, "text"),
    ("preview_image",             50, "url"),
    ("resource_urls_count",       8,  "int"),
    ("enrichment_summary",        60, "wrap"),
    ("enrichment_transcription",  60, "wrap"),
    ("enrichment_analysis",       60, "wrap"),
    ("enrichment_ugc_script",     60, "wrap"),
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

GOLD_FILL = PatternFill(fill_type="solid", fgColor="FFD700")
SILVER_FILL = PatternFill(fill_type="solid", fgColor="C0C0C0")

WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
URL_FONT = Font(color="0563C1", underline="single")

NO_ENRICHMENT_MARKER = "[no enrichment]"


def row_from_ad(ad: dict) -> list:
    """Project an ad record into the 20-column row tuple."""
    enr = ad.get("enrichment") if isinstance(ad.get("enrichment"), dict) else None
    if enr is None:
        summary = transcription = analysis = ugc = NO_ENRICHMENT_MARKER
    else:
        summary = enr.get("summary") or ""
        transcription = enr.get("transcription") or ""
        analysis = enr.get("analysis") or ""
        ugc = enr.get("ugc_script") or ""

    resource_urls = ad.get("resource_urls")
    resource_count = len(resource_urls) if isinstance(resource_urls, list) else 0

    source_keywords = ad.get("_source_keywords") or []
    overlap_count = len(source_keywords) if isinstance(source_keywords, list) else 0

    return [
        ad.get("_tier", ""),
        ad.get("advertiser_name", ""),
        ad.get("platform", ""),
        int(ad.get("days_count", 0) or 0),
        int(ad.get("heat", 0) or 0),
        int(ad.get("impression", 0) or 0),
        int(ad.get("estimated_spend", 0) or 0),
        int(ad.get("video_duration", 0) or 0),
        ad.get("call_to_action", ""),
        ad.get("_source_keyword", ""),
        ad.get("_source_bucket", ""),
        overlap_count,
        int(ad.get("ads_type", 0) or 0),
        str(ad.get("ads_promote_type", "")),
        ad.get("preview_img_url", ""),
        resource_count,
        summary,
        transcription,
        analysis,
        ugc,
    ]


def sort_ads(ads: list) -> list:
    """Gold first, silver after; within each tier descending by heat."""
    tier_rank = {"gold": 0, "silver": 1}

    def key(ad: dict):
        return (
            tier_rank.get(ad.get("_tier"), 99),
            -(int(ad.get("heat", 0) or 0)),
        )

    return sorted(ads, key=key)


def main() -> int:
    if not INPUT_PATH.exists():
        sys.stderr.write(f"ERROR: input not found at {INPUT_PATH}\n")
        return 1

    with INPUT_PATH.open("r", encoding="utf-8") as fh:
        doc = json.load(fh)

    ads = doc.get("ads")
    if not isinstance(ads, list) or not ads:
        sys.stderr.write("ERROR: input has no 'ads' array or it is empty.\n")
        return 1

    ads = sort_ads(ads)
    print(f"[init] {len(ads)} ads loaded and sorted (gold -> silver, heat desc)")

    wb = Workbook()
    ws = wb.active
    ws.title = "ads"

    # Header
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for col_idx, _ in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # Column widths
    for col_idx, (_, width, _kind) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for ad in ads:
        ws.append(row_from_ad(ad))

    # Per-cell styling pass (tier fill, hyperlinks, wrap)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx, (_header, _width, kind) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if kind == "tier":
                val = (cell.value or "").lower()
                if val == "gold":
                    cell.fill = GOLD_FILL
                elif val == "silver":
                    cell.fill = SILVER_FILL
            elif kind == "wrap":
                cell.alignment = WRAP_ALIGN
            elif kind == "url":
                url = cell.value
                if isinstance(url, str) and url.strip():
                    cell.hyperlink = url
                    cell.font = URL_FONT

    # Freeze + filter
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    wb.save(OUTPUT_PATH)

    size_kb = OUTPUT_PATH.stat().st_size / 1024.0
    print(f"OUTPUT: {OUTPUT_PATH}")
    print(f"  size: {size_kb:.1f} KB")
    print(f"  rows: {ws.max_row} (1 header + {ws.max_row - 1} data)")
    print(f"  cols: {ws.max_column}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
